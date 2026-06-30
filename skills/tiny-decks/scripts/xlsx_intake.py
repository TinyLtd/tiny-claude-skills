#!/usr/bin/env python3
"""
xlsx_intake.py — deterministic Excel table extraction + classification.

PHILOSOPHY: extraction, NOT mapping
-----------------------------------
This script does ONE job well: open one or more .xlsx workbooks, find every
data table across all sheets, resolve merged cells, and make a *best-guess*
classification of each table (p&l, trend, by-company, fund, cap table, balance
sheet, cash flow, kpi, or unknown). It deliberately does NOT try to map those
tables onto the final deck model in references/data-schema.md.

Why: arbitrary finance workbooks are too messy and too varied for a script to
map reliably. That mapping is Claude's job — Claude reads this script's output
(both the machine JSON and the human summary printed to stdout), reasons about
which block is which deck section, and asks the user to confirm anything that
is ambiguous, "unknown", or low-confidence. So this script optimizes for
*recall* (find every table, never crash) and *legible signal* (clear ranges,
headers, samples, a confidence-scored guess), not for a perfect final answer.

WHAT IT EMITS
-------------
1. Machine JSON (--out, default /tmp/xlsx_blocks.json):
   {
     "files": [
       {
         "path": "<abs path>",
         "warnings": ["..."],
         "sheets": [
           {
             "name": "Model",
             "blocks": [
               {
                 "range": "B3:H17",
                 "rows": 14, "cols": 7,
                 "classification": "pnl",
                 "confidence": 0.62,
                 "headers": ["Company", "Revenue", "COGS", ...],
                 "sample": [["Dribbble", 1200, 400, ...], ...]   # up to ~8 rows
               }
             ]
           }
         ]
       }
     ]
   }

2. Human summary to stdout: per file -> per sheet -> one line per block:
       [B3:H17] PNL (0.62) — headers: Company, Revenue, COGS (14 rows)
   ending with a hint telling Claude to map blocks against the deck schema and
   confirm unknown/low-confidence blocks with the user.

USAGE
-----
    python xlsx_intake.py <file1.xlsx> [file2.xlsx ...] [--out blocks.json]

DEPENDENCIES
------------
openpyxl only (3.1.5 tested). Does NOT import pandas (may be missing locally).

READ-ONLY MODE TRADEOFF (important)
-----------------------------------
The task spec suggested load_workbook(..., read_only=True). In practice
read_only=True makes BOTH ws.tables (Excel named tables / ListObjects) AND
ws.merged_cells unavailable — the two cleanest structural signals. Since this
script's whole value is high-recall structural extraction, we default to
NON-read_only load_workbook(data_only=True), which makes tables + merged cells
reliable, at the cost of higher memory (whole workbook in RAM). For the board
packs this skill targets (a handful of sheets, thousands of cells) that is
fine. Pass --read-only to force the low-memory streaming path for very large
files; classification still works but named-table and merged-cell resolution
degrade gracefully (we fall back to blank-gap block scanning + raw reads).
"""
import json
import re
import sys

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except Exception as e:  # pragma: no cover
    print(f"FATAL: openpyxl required ({e})", file=sys.stderr)
    sys.exit(2)


# --------------------------------------------------------------------------
# low-level helpers
# --------------------------------------------------------------------------
def _num(v):
    """Coerce a cell value to a number when it clearly is one.

    Native numbers pass through. Strings like '$1,234', '(500)', '27%' are
    coerced ('(...)' -> negative). Anything else returns the original value
    unchanged (we keep text labels as-is for headers/line items).
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if not isinstance(v, str):
        return v
    t = v.strip()
    if t == "" or t in ("-", "—", "–", "n/a", "N/A"):
        return None if t == "" else v
    raw = t
    neg = False
    if t.startswith("(") and t.endswith(")"):
        neg = True
        t = t[1:-1]
    t = t.replace(",", "").replace("$", "").replace("%", "").strip()
    if not re.fullmatch(r"-?\d+(\.\d+)?", t):
        return raw  # not numeric — leave the original string
    f = float(t)
    if f == int(f):
        f = int(f)
    return -f if neg else f


def _txt(v):
    return "" if v is None else str(v).strip()


def _is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


# period-label detectors -----------------------------------------------------
_MONTHS = ("jan", "feb", "mar", "apr", "may", "jun",
           "jul", "aug", "sep", "oct", "nov", "dec")
_RE_MONTH = re.compile(r"\b(" + "|".join(_MONTHS) + r")\w*\b", re.I)
_RE_QTR = re.compile(r"\bq[1-4]\b", re.I)
_RE_YEAR = re.compile(r"\b(19|20)\d{2}\b")


def _is_period(v):
    s = _txt(v)
    if not s:
        return False
    if _RE_MONTH.search(s) or _RE_QTR.search(s) or _RE_YEAR.search(s):
        return True
    # bare 4-digit year as a number
    if isinstance(v, (int, float)) and 1900 <= int(v) <= 2099 and float(v) == int(v):
        return True
    return False


# --------------------------------------------------------------------------
# merged-cell resolution
# --------------------------------------------------------------------------
def _build_merge_map(ws):
    """Return {(row,col): top-left value} for every merged range, propagating
    the anchor value across the whole range. Empty/unavailable -> {}."""
    out = {}
    try:
        ranges = list(ws.merged_cells.ranges)
    except Exception:
        return out
    for mr in ranges:
        try:
            anchor = ws.cell(row=mr.min_row, column=mr.min_col).value
        except Exception:
            continue
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                out[(r, c)] = anchor
    return out


def _grid(ws, merge_map):
    """Materialize the used range into a dense list-of-lists of values with
    merged cells propagated. Returns (grid, min_row, min_col) where grid[i][j]
    is the value at sheet row=min_row+i, col=min_col+j. Empty sheet -> ([],..)."""
    rows = []
    min_r = min_c = None
    max_r = max_c = 0
    # First pass: collect cells, track bounds.
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            r, c = cell.row, cell.column
            val = cell.value
            mv = merge_map.get((r, c))
            if mv is not None and _is_blank(val):
                val = mv
            if not _is_blank(val):
                cells[(r, c)] = val
                min_r = r if min_r is None else min(min_r, r)
                min_c = c if min_c is None else min(min_c, c)
                max_r = max(max_r, r)
                max_c = max(max_c, c)
    if min_r is None:
        return [], None, None
    grid = []
    for r in range(min_r, max_r + 1):
        grid.append([cells.get((r, c)) for c in range(min_c, max_c + 1)])
    return grid, min_r, min_c


# --------------------------------------------------------------------------
# block detection (blank-row + blank-column separated rectangles)
# --------------------------------------------------------------------------
def _find_blocks(grid):
    """Flood-fill contiguous non-blank cells where adjacency is 4-directional.
    A fully-blank row OR column naturally severs neighbours. Returns a list of
    (r0, c0, r1, c1) bounding boxes (grid-local, inclusive)."""
    if not grid:
        return []
    nrows, ncols = len(grid), len(grid[0])
    seen = [[False] * ncols for _ in range(nrows)]
    boxes = []
    for i in range(nrows):
        for j in range(ncols):
            if seen[i][j] or _is_blank(grid[i][j]):
                continue
            # BFS flood fill
            stack = [(i, j)]
            seen[i][j] = True
            r0 = r1 = i
            c0 = c1 = j
            while stack:
                ci, cj = stack.pop()
                r0, r1 = min(r0, ci), max(r1, ci)
                c0, c1 = min(c0, cj), max(c1, cj)
                for di, dj in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                    ni, nj = ci + di, cj + dj
                    if 0 <= ni < nrows and 0 <= nj < ncols \
                            and not seen[ni][nj] and not _is_blank(grid[ni][nj]):
                        seen[ni][nj] = True
                        stack.append((ni, nj))
            boxes.append((r0, c0, r1, c1))
    # Merge boxes whose bounding rectangles overlap/touch (handles ragged tables
    # with internal blank cells that flood-fill split into pieces).
    boxes = _merge_overlapping(boxes)
    return boxes


def _merge_overlapping(boxes):
    changed = True
    while changed:
        changed = False
        out = []
        while boxes:
            b = boxes.pop()
            r0, c0, r1, c1 = b
            merged = True
            while merged:
                merged = False
                rest = []
                for o in boxes:
                    # touch/overlap test with 0-gap tolerance
                    if not (o[0] > r1 + 1 or o[2] < r0 - 1 or
                            o[1] > c1 + 1 or o[3] < c0 - 1):
                        r0, c0 = min(r0, o[0]), min(c0, o[1])
                        r1, c1 = max(r1, o[2]), max(c1, o[3])
                        merged = True
                        changed = True
                    else:
                        rest.append(o)
                boxes = rest
            out.append((r0, c0, r1, c1))
        boxes = out
    return boxes


# --------------------------------------------------------------------------
# classification
# --------------------------------------------------------------------------
# keyword sets per type (lowercased substring match against header + col-A text)
_KW = {
    "pnl": ("revenue", "cogs", "cost of goods", "gross profit", "gross margin",
            "ebitda", "net income", "operating income", "opex", "margin",
            "s&m", "sales & marketing", "g&a", "p&l", "profit"),
    "by_company": ("company", "subsidiary", "portfolio company", "brand",
                   "business unit", "segment"),
    "fund": ("holding", "holdings", "ownership", "fund", "lp interest",
             "stake", "% owned", "portfolio"),
    "captable": ("shares", "share class", "fully diluted", "fully-diluted",
                 "investor", "option pool", "preferred", "common", "cap table",
                 "ownership %", "% ownership"),
    "balance_sheet": ("assets", "liabilities", "total equity", "cash and",
                      "accounts receivable", "accounts payable", "retained earnings",
                      "balance sheet", "current assets", "current liabilities",
                      "goodwill", "inventory"),
    "fcf": ("free cash flow", "cash flow", "interest", "taxes", "capex",
            "capital expenditure", "fcf", "working capital", "depreciation"),
}


def _cells_text(rows):
    return [_txt(c).lower() for row in rows for c in row if not _is_blank(c)]


def _classify(headers, data_rows, all_rows):
    """Return (label, confidence 0..1, reasons). Score = keyword hits + shape."""
    header_txt = " | ".join(_txt(h).lower() for h in headers)
    cola = [_txt(r[0]).lower() for r in all_rows if r]  # left column = line items
    blob = " ".join(_cells_text(all_rows))
    scores = {}

    # --- keyword scoring ---------------------------------------------------
    for label, kws in _KW.items():
        hits = sum(1 for kw in kws if kw in blob)
        if hits:
            scores[label] = scores.get(label, 0.0) + hits * 1.0

    # --- shape signals -----------------------------------------------------
    # trend: a row OR column of consecutive period labels
    period_in_header = sum(1 for h in headers if _is_period(h))
    # left column periods (vertical trend)
    period_in_cola = sum(1 for r in all_rows if r and _is_period(r[0]))
    max_period_run = max(period_in_header, period_in_cola)
    if max_period_run >= 3:
        scores["trend"] = scores.get("trend", 0.0) + max_period_run * 0.6

    # by_company: company-ish names down col A + revenue/ebitda across header
    if ("revenue" in header_txt or "ebitda" in header_txt or "rev" in header_txt) \
            and len([x for x in cola if x]) >= 3:
        scores["by_company"] = scores.get("by_company", 0.0) + 1.5

    # pnl: classic line items down col A
    pnl_lineitems = ("revenue", "cogs", "gross", "ebitda", "net income",
                     "operating", "margin")
    li_hits = sum(1 for x in cola if any(li in x for li in pnl_lineitems))
    if li_hits >= 2:
        scores["pnl"] = scores.get("pnl", 0.0) + li_hits * 0.8

    # fund: ownership/holding columns are a strong fund signal — boost it above
    # a generic rev/ebitda P&L read when an ownership column is present.
    if ("ownership" in header_txt or "% owned" in header_txt or "stake" in header_txt
            or "holding" in header_txt):
        scores["fund"] = scores.get("fund", 0.0) + 2.0
        # ownership column shouldn't read as a plain P&L
        if "pnl" in scores and "holding" in header_txt:
            scores["pnl"] = scores["pnl"] * 0.5

    # captable: ownership %-heavy
    if "%" in blob and any(k in blob for k in ("shares", "investor", "diluted", "preferred")):
        scores["captable"] = scores.get("captable", 0.0) + 1.5

    # kpi: very small block, few cells, big labeled numbers (label,value pairs).
    n_data = len(data_rows)
    n_cells = sum(1 for r in all_rows for c in r if not _is_blank(c))
    looks_kpi = (ncols := max((len(r) for r in all_rows), default=0)) <= 3 and len(all_rows) <= 6
    if looks_kpi and (not scores or max(scores.values()) < 1.5):
        # value column present?
        has_vals = any(isinstance(_num(r[-1]), (int, float)) for r in all_rows if len(r) >= 2)
        if has_vals or n_cells <= 8:
            scores["kpi"] = scores.get("kpi", 0.0) + 1.5

    if not scores:
        return "unknown", 0.15, ["no keyword or shape signal"]

    # pick winner
    best = max(scores, key=scores.get)
    top = scores[best]
    total = sum(scores.values())
    # confidence: dominance of the winner, squashed into a sane band
    dominance = top / total if total else 0.0
    raw = 0.35 + 0.5 * dominance + min(top, 5) * 0.04
    conf = round(min(0.97, raw), 2)
    if top < 1.0:
        conf = round(min(conf, 0.4), 2)
    reasons = [f"{k}={round(v, 1)}" for k, v in sorted(scores.items(), key=lambda x: -x[1])]
    return best, conf, reasons


# --------------------------------------------------------------------------
# block -> structured record
# --------------------------------------------------------------------------
def _pick_header_row(block_rows):
    """Best-guess header row = first row that is mostly text, else the modal-
    width row. Returns index into block_rows."""
    widths = [sum(1 for c in r if not _is_blank(c)) for r in block_rows]
    for i, r in enumerate(block_rows):
        non_blank = [c for c in r if not _is_blank(c)]
        if not non_blank:
            continue
        text_frac = sum(1 for c in non_blank if isinstance(c, str)) / len(non_blank)
        if text_frac >= 0.6 and len(non_blank) >= 2:
            return i
    # fallback: modal width row (first one achieving the max width)
    if widths:
        mx = max(widths)
        return widths.index(mx)
    return 0


def _range_str(min_r, min_c, box):
    r0, c0, r1, c1 = box
    a = f"{get_column_letter(min_c + c0)}{min_r + r0}"
    b = f"{get_column_letter(min_c + c1)}{min_r + r1}"
    return f"{a}:{b}"


def _block_record(grid, min_r, min_c, box, warn):
    r0, c0, r1, c1 = box
    block_rows = [grid[i][c0:c1 + 1] for i in range(r0, r1 + 1)]
    nrows = len(block_rows)
    ncols = (c1 - c0 + 1)

    hidx = _pick_header_row(block_rows)
    headers = [(_txt(c) if c is not None else "") for c in block_rows[hidx]]
    data_rows = block_rows[hidx + 1:]
    # sample up to 8 data rows, coercing obvious numeric strings
    sample = []
    for r in data_rows[:8]:
        sample.append([_num(c) for c in r])

    label, conf, reasons = _classify(headers, data_rows, block_rows)

    return {
        "range": _range_str(min_r, min_c, box),
        "rows": nrows,
        "cols": ncols,
        "classification": label,
        "confidence": conf,
        "headers": headers,
        "sample": sample,
        "_reasons": reasons,
    }


# --------------------------------------------------------------------------
# named-table (ListObject) path
# --------------------------------------------------------------------------
def _table_blocks(ws, merge_map, warn):
    """Read Excel named tables directly. Returns list of block records or []."""
    out = []
    try:
        tables = dict(ws.tables)
    except Exception:
        return []
    for name, tbl in tables.items():
        try:
            ref = tbl.ref  # e.g. "B3:H17"
            cells = ws[ref]
            if not isinstance(cells, tuple):
                cells = (cells,)
            block_rows = []
            for row in cells:
                rr = []
                for cell in (row if isinstance(row, tuple) else (row,)):
                    v = cell.value
                    mv = merge_map.get((cell.row, cell.column))
                    if mv is not None and _is_blank(v):
                        v = mv
                    rr.append(v)
                block_rows.append(rr)
            if not block_rows:
                continue
            headers = [_txt(c) for c in block_rows[0]]
            data_rows = block_rows[1:]
            sample = [[_num(c) for c in r] for r in data_rows[:8]]
            label, conf, reasons = _classify(headers, data_rows, block_rows)
            out.append({
                "range": ref,
                "rows": len(block_rows),
                "cols": len(block_rows[0]) if block_rows else 0,
                "classification": label,
                "confidence": conf,
                "headers": headers,
                "sample": sample,
                "_reasons": reasons + [f"named-table:{name}"],
            })
        except Exception as e:
            warn(f"named table '{name}' failed: {e}")
    return out


# --------------------------------------------------------------------------
# per-sheet / per-file orchestration
# --------------------------------------------------------------------------
def _process_sheet(ws, warn):
    merge_map = _build_merge_map(ws)

    # 1. named tables first (cleanest signal)
    blocks = _table_blocks(ws, merge_map, warn)
    if blocks:
        return blocks

    # 2. fall back to blank-gap block scanning
    grid, min_r, min_c = _grid(ws, merge_map)
    if not grid:
        return []
    boxes = _find_blocks(grid)
    out = []
    for box in boxes:
        try:
            # skip 1-cell noise
            r0, c0, r1, c1 = box
            if (r1 - r0 + 1) * (c1 - c0 + 1) < 2:
                continue
            out.append(_block_record(grid, min_r, min_c, box, warn))
        except Exception as e:
            warn(f"block {box} failed: {e}")
    # order top-to-bottom, left-to-right by their starting cell
    out.sort(key=lambda b: b["range"])
    return out


def _process_file(path, force_read_only):
    warnings = []
    result = {"path": path, "warnings": warnings, "sheets": []}
    try:
        wb = load_workbook(path, data_only=True, read_only=force_read_only)
    except Exception as e:
        warnings.append(f"could not open workbook: {e}")
        return result

    any_value = False
    for ws in wb.worksheets:
        sheet_rec = {"name": ws.title, "blocks": []}
        try:
            blocks = _process_sheet(ws, warnings.append)
            sheet_rec["blocks"] = blocks
            if blocks:
                any_value = True
        except Exception as e:
            warnings.append(f"sheet '{ws.title}' failed: {e}")
        result["sheets"].append(sheet_rec)

    if not any_value:
        warnings.append("no cached values found across all sheets — "
                        "re-save in Excel so cached values are written")
    try:
        wb.close()
    except Exception:
        pass
    return result


# --------------------------------------------------------------------------
# summary printer
# --------------------------------------------------------------------------
def _print_summary(files):
    lines = []
    for f in files:
        lines.append(f"\n=== {f['path']} ===")
        if f["warnings"]:
            for w in f["warnings"][:8]:
                lines.append(f"  ! {w}")
        if not f["sheets"]:
            lines.append("  (no sheets)")
        for sh in f["sheets"]:
            nb = len(sh["blocks"])
            lines.append(f"  Sheet '{sh['name']}' — {nb} block(s)")
            for b in sh["blocks"]:
                hdrs = ", ".join(h for h in b["headers"] if h)[:80] or "(none)"
                lines.append(
                    f"    [{b['range']}] {b['classification'].upper()} "
                    f"({b['confidence']}) — headers: {hdrs} ({b['rows']} rows)"
                )
    lines.append(
        "\nClaude: map each block to the deck model in references/data-schema.md; "
        "ask the user to confirm any 'unknown' or low-confidence blocks before building."
    )
    print("\n".join(lines))


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------
def main(argv):
    args = list(argv)
    out_path = "/tmp/xlsx_blocks.json"
    force_read_only = False
    if "--out" in args:
        i = args.index("--out")
        out_path = args[i + 1]
        del args[i:i + 2]
    if "--read-only" in args:
        force_read_only = True
        args.remove("--read-only")
    paths = args
    if not paths:
        print(__doc__)
        return 1

    files = [_process_file(p, force_read_only) for p in paths]
    payload = {"files": files}

    # strip internal _reasons from JSON? keep them — useful for Claude.
    with open(out_path, "w") as fh:
        json.dump(payload, fh, indent=2, default=str)

    _print_summary(files)
    print(f"\n[xlsx_intake] JSON -> {out_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
